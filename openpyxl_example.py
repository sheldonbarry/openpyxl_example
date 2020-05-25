import openpyxl

# Create a new workbook instance and select the active sheet
wb = openpyxl.Workbook()
sheet = wb.active

# dict of superhero data
superheroes = {
	'Captain America':'Steve Rogers', 
	'Iron Man':'Tony Stark', 
	'Spiderman':'Peter Parker', 
	'Hulk':'Bruce Banner', 
	'Superman':'Clark Kent', 
	'Batman':'Bruce Wayne', 
	'Wonder Woman':'Princess Diana'}

# Write 'Superhero' heading to cell A1 and set font style to bold
sheet['A1'].value = 'Superhero'
sheet['A1'].font = openpyxl.styles.Font(bold=True)

# Write 'Real Name' heading to cell A1 and set font style to bold
sheet['B1'].value = 'Real Name'
sheet['B1'].font = openpyxl.styles.Font(bold=True)

# Starting at row 2, write data to sheet (alphabetically by key)
row = 2
for key in sorted(superheroes.keys()):
	sheet.cell(column=1, row=row, value=key)
	sheet.cell(column=2, row=row, value=superheroes[key])
	row += 1

# Save the file
myFile = 'heroes.xlsx'
wb.save(filename = myFile)

# open active sheet in the workbook and create new workbook and sheet objects
wb1 = openpyxl.load_workbook(myFile)
sheet1 = wb1.active

# Read back data from the worksheet by cell reference
print ('{}\t\t{}'.format(sheet1['A3'].value, sheet1['B3'].value))
print

# Read back all the rows in the worksheet
sheet_rows = sheet1.iter_rows()

# (warning - iter_rows starts counting at 0)
for row in sheet_rows:
	print('{}{}{}'.format(row[0].value, ' '*(20-len(row[0].value)), row[1].value))
